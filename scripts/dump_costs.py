#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Costos USD por SKU (mejor y último) + material, desde tablas mensuales + historial."""
import openpyxl, glob, json, re, datetime

import os as _os
_REPO_DATOS = _os.path.join(_os.path.dirname(_os.path.dirname(_os.path.abspath(__file__))), 'datos')
_REPO = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
DL = _os.path.expanduser('~/Downloads')
if not _os.path.isdir(DL + '/TABLAS DE PRECIOS') and _os.path.isdir(_REPO_DATOS + '/TABLAS DE PRECIOS'):
    DL = _REPO_DATOS   # en otra compu: usa los Excel guardados en el repo
# Rutas portables (igual que build_ordenes.py): OC_SCRATCH o <REPO>/.scratch
SCRATCH = _os.environ.get('OC_SCRATCH') or _os.path.join(_REPO, '.scratch')
_os.makedirs(SCRATCH, exist_ok=True)

def norm_sku(v):
    s = str(v or '').strip()
    if s.endswith('.0'): s = s[:-2]
    return s

def to_num(v):
    if v is None or isinstance(v, (int, float)): return v
    s = str(v).strip().replace('$', '').replace(',', '').strip()
    try: return float(s)
    except ValueError: return None

def parse_fecha(v, ctx=''):
    if isinstance(v, datetime.datetime): return v.date()
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', str(v).strip())
    if m: return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    if 'MAYO 2025' in ctx: return datetime.date(2025, 5, 1)
    return None

files = sorted(glob.glob(DL + '/TABLAS DE PRECIOS/*.xlsx'))
files = [f for f in files if not f.endswith('2026-02_FEBRERO_2026.xlsx')]
costs = {}
for f in files:
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb['TODAS LAS ORDENES']
    it = ws.iter_rows(values_only=True); next(it)
    for r in it:
        if not r[0] or r[3] is None: continue
        fecha = parse_fecha(r[2], str(r[1] or ''))
        cu = to_num(r[7])
        if cu is None: continue
        sku = norm_sku(r[3])
        c = costs.setdefault(sku, {'prov': str(r[0]).strip(), 'best': cu, 'last': cu,
                                   'last_fecha': fecha.isoformat() if fecha else ''})
        if cu < c['best']: c['best'] = cu
        if fecha and fecha.isoformat() >= c['last_fecha']:
            c['last'] = cu; c['last_fecha'] = fecha.isoformat()
    wb.close()

hwb = openpyxl.load_workbook(DL + '/HISTORIA PRECIOS NO ACTUALIZADO.xlsx', data_only=True)
hist_map = {'CYNTHIA CAO': 'CYNTHIA', 'NANCY VIP': 'NANCY', 'HAIFENG': 'HAIFENG', 'ZOEY': 'ZOEY', 'DINA DU': 'DINA DU'}
hist_only = {'CYNTHIA CAO': [('CYN070726', '2026-07-07')], 'ZOEY': [('Dic\n2025', '2025-12-01')],
             'DINA DU': [('Dic\n2024', '2024-12-01'), ('Ene\n2025', '2025-01-01')]}
material = {}
cyn_jul = []   # orden CYNTHIA 07/07 para transito
for sheet, prov in hist_map.items():
    ws = hwb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h) if h is not None else '' for h in rows[3]]
    i_mat = header.index('MATERIAL')
    onlyc = []
    for (htxt, fdate) in hist_only.get(sheet, []):
        for i, h in enumerate(header):
            if htxt in h: onlyc.append((i, fdate)); break
    for r in rows[4:]:
        if r[0] is None: continue
        sku = norm_sku(r[0])
        if r[i_mat] not in (None, '—'): material[sku] = str(r[i_mat]).strip()
        for i, fdate in onlyc:
            v = r[i]
            if isinstance(v, (int, float)):
                c = costs.setdefault(sku, {'prov': prov, 'best': float(v), 'last': float(v), 'last_fecha': fdate})
                if v < c['best']: c['best'] = float(v)
                if fdate >= c['last_fecha']: c['last'] = float(v); c['last_fecha'] = fdate
                if sheet == 'CYNTHIA CAO':
                    vt = str(r[3] or '').replace('\n', ' ').strip()
                    cyn_jul.append({'title': str(r[1] or '').strip(), 'variant': vt, 'sku': sku, 'qty': 0, 'arrived': False})

json.dump({'costs': costs, 'material': material}, open(SCRATCH + '/costs.json', 'w'))

# transit base: NANCY 09/07 (tabla julio) + CYNTHIA 07/07 (historial)
wb = openpyxl.load_workbook(DL + '/TABLAS DE PRECIOS/2026-07_JULIO_2026.xlsx', read_only=True, data_only=True)
ws = wb['TODAS LAS ORDENES']
it = ws.iter_rows(values_only=True); next(it)
nancy = []
for r in it:
    if not r[0]: continue
    nancy.append({'title': str(r[4]).strip(), 'variant': str(r[5] or '').strip(),
                  'sku': norm_sku(r[3]), 'qty': int(r[6] or 0), 'arrived': False})
wb.close()
base = [{'name': 'NANCY 09/07/2026 (precargada)', 'date': '2026-07-09', 'base': 1, 'lines': nancy},
        {'name': 'CYNTHIA 07/07/2026 (precargada, sin cantidades)', 'date': '2026-07-07', 'base': 1, 'lines': cyn_jul}]
# OJO: este "base" se arma SOLO con NANCY 09/07 + CYNTHIA 07/07 (sin cantidades) y NO incluye
# HAIFENG 11/07. El transit_base.json bueno (las 3 ordenes, con cantidades) se armo a mano desde
# los Excel de Eduardo y vive en scripts/transit_base.json. Por eso aqui se escribe a un archivo
# APARTE: sobrescribir transit_base.json borraria el "en camino" y haria sobre-pedir mercancia.
json.dump(base, open(SCRATCH + '/transit_base_auto.json', 'w'), ensure_ascii=False)
print('costos:', len(costs), '| materiales:', len(material), '| transito NANCY:', len(nancy), 'CYNTHIA:', len(cyn_jul))
